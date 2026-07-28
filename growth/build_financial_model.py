#!/usr/bin/env python3
"""
Taruu financial model generator.
Emits financial-model.xlsx + prints key figures for FINANCIAL-MODEL.md.

Locked decisions (2026-06-29):
  - MEMBERSHIP model: only the FIRST vote of a calendar month costs money — ₪6.
    Every other vote that month is FREE. Charged once/month per active member,
    card-on-file (Green Invoice). Members who don't vote in a month pay ₪0.
  - The ₪6 splits: ₪2.10 → civic treasury POOL (per member/month), ₪3.90 → platform.
    Treasury is now a MONTHLY POOL allocated to the decisions executed that month,
    NOT a per-vote ₪2.10. (Free votes can't fund per-vote treasury.)
  - Create vote: gross ₪50, 100% platform (unchanged).
  - Rail: Green Invoice (Prime), card-on-file. GI fee = 1.4% + ₪1.2 + ₪0.15 receipt.
  - Strategic bet: free voting expands the funnel (more members + more creates).
  - Money execution in-house → treasury/ops salary is real opex (estimate, flagged).
  - Target: ₪30–45k/mo combined take-home (midpoint ₪37.5k).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

USD_ILS = 3.7
GI_PCT = 0.014
GI_FIXED = 1.20
GI_RECEIPT = 0.15
GI_PLAN_MO = 155.0                       # Prime plan

MEMBER_FEE = 6.0                         # first vote of the calendar month
TREASURY_PER_MEMBER = 2.10               # to the monthly civic pool
PLATFORM_SHARE = MEMBER_FEE - TREASURY_PER_MEMBER  # ₪3.90
CREATE_GROSS = 50.0
AVG_VOTES_PER_MEMBER = 8                 # assumption — for participation-volume context only

FIXED_COSTS = 1200.0 + GI_PLAN_MO
TAKEHOME_TARGET = 37500.0
EFF_TAX = 0.27
GROSS_PROFIT_TARGET = round(TAKEHOME_TARGET / (1 - EFF_TAX) + FIXED_COSTS)

def gi_fee(amount):
    return GI_PCT * amount + GI_FIXED + GI_RECEIPT

# ---- Unit economics ----
member_fee_cost = gi_fee(MEMBER_FEE)
member_net = PLATFORM_SHARE - member_fee_cost          # platform net per active member/month
create_fee = gi_fee(CREATE_GROSS)
create_net = CREATE_GROSS - create_fee

# ---- Scenarios (driver = active paying members/month + paid creates/month) ----
def scenario(name, ambassadors, members, creates, treasury_ops, extra):
    rev_members = members * member_net
    rev_creates = creates * create_net
    gross = rev_members + rev_creates
    costs = FIXED_COSTS + treasury_ops + extra
    gp = gross - costs
    takehome = max(0, gp) * (1 - EFF_TAX)
    return {
        "name": name, "ambassadors": ambassadors, "members": members, "creates": creates,
        "votes_ctx": members * AVG_VOTES_PER_MEMBER,
        "rev_members": round(rev_members), "rev_creates": round(rev_creates),
        "treasury_ops": treasury_ops, "civic_pool": round(members * TREASURY_PER_MEMBER),
        "gross": round(gross), "costs": round(costs), "gross_profit": round(gp),
        "est_takehome": round(takehome),
    }

scenarios = [
    # name, ambassadors, paying members/mo, paid creates/mo, treasury-ops, extra
    scenario("Break-even",      8,   250,   30,    0,  200),
    scenario("Lean (ramp)",    20,  1500,  150, 2500,  500),
    scenario("Target (band)",  60,  8000,  900, 5000, 1200),   # the growth bet: ~1.5-2x old engagement
    scenario("Scale",         150, 30000, 3000, 8000, 2500),
]

# creates-only break-even (cover fixed costs, zero salary)
be_creates_only = FIXED_COSTS / create_net
target_creates_only = GROSS_PROFIT_TARGET / create_net

# ---------- workbook ----------
wb = Workbook()
H = Font(bold=True, color="FFFFFF"); HEAD = PatternFill("solid", fgColor="14110E")
GREEN = PatternFill("solid", fgColor="1F7A3D"); SUB = Font(bold=True)

def style_header(ws, row, ncols, fill=HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c); cell.font = H; cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def autosize(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 40)

ws = wb.active; ws.title = "Assumptions"
ws["A1"] = "TARUU — FINANCIAL MODEL"; ws["A1"].font = Font(bold=True, size=16)
ws["A2"] = "Updated 2026-06-29 · ₪ · MEMBERSHIP model (first vote/month ₪6, rest free)"
rows = [
    ("Input", "Value", "Note"),
    ("Member fee / month", MEMBER_FEE, "first vote of the calendar month; rest free"),
    ("→ Treasury share", TREASURY_PER_MEMBER, "to monthly civic pool / member"),
    ("→ Platform share", PLATFORM_SHARE, "before GI fee"),
    ("GI fee on ₪6", round(member_fee_cost, 2), "1.4% + ₪1.2 + ₪0.15 receipt"),
    ("Platform NET / member / mo", round(member_net, 2), "after GI fee"),
    ("Create gross / NET", f"{CREATE_GROSS} / {round(create_net,2)}", "100% platform"),
    ("Avg votes / member / mo", AVG_VOTES_PER_MEMBER, "assumption — funnel context only"),
    ("GI plan / mo", GI_PLAN_MO, "Prime"),
    ("Fixed costs / mo", FIXED_COSTS, "infra + tooling + GI Prime"),
    ("Take-home target", TAKEHOME_TARGET, "midpoint ₪30–45k"),
    ("Effective tax+BL", f"{int(EFF_TAX*100)}%", "2x osek murshe"),
    ("→ Gross-profit target", GROSS_PROFIT_TARGET, "to net the take-home"),
]
for i, r in enumerate(rows, start=4):
    for j, v in enumerate(r, start=1): ws.cell(row=i, column=j, value=v)
style_header(ws, 4, 3); autosize(ws)

ws = wb.create_sheet("Scenarios")
ws["A1"] = "MONTHLY SCENARIOS (membership + create-led)"; ws["A1"].font = Font(bold=True, size=13)
cols = ["Scenario", "Ambassadors", "Paying members", "Votes (ctx)", "Paid creates",
        "Member rev", "Create rev", "Treasury ops", "Gross rev", "Costs",
        "Gross profit", "Take-home", "Civic pool"]
for j, c in enumerate(cols, start=1): ws.cell(row=3, column=j, value=c)
style_header(ws, 3, len(cols))
for i, s in enumerate(scenarios, start=4):
    vals = [s["name"], s["ambassadors"], s["members"], s["votes_ctx"], s["creates"],
            s["rev_members"], s["rev_creates"], s["treasury_ops"], s["gross"], s["costs"],
            s["gross_profit"], s["est_takehome"], s["civic_pool"]]
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=j, value=v)
        if s["name"].startswith("Target"): cell.fill = PatternFill("solid", fgColor="FCE9C8")
autosize(ws)

wb.save("growth/financial-model.xlsx")

# ---- print for the markdown ----
print("=== UNIT ECONOMICS (membership model) ===")
print(f"  Member ₪6/mo: GI fee ₪{member_fee_cost:.2f} | treasury ₪{TREASURY_PER_MEMBER} | platform NET ₪{member_net:.2f}/member/mo")
print(f"  Create ₪50: GI fee ₪{create_fee:.2f} -> platform NET ₪{create_net:.2f}")
print("=== BREAK-EVEN (creates-only) ===")
print(f"  cover fixed costs: {be_creates_only:.1f} creates/mo ({be_creates_only/30:.1f}/day)")
print(f"  GP target ₪{GROSS_PROFIT_TARGET}: {target_creates_only:.0f} creates/mo (if creates carried it alone)")
print("=== SCENARIOS ===")
for s in scenarios:
    print(f"  {s['name']:<14} amb={s['ambassadors']:>3} members={s['members']:>6} creates={s['creates']:>4} | "
          f"gross ₪{s['gross']:>6} GP ₪{s['gross_profit']:>6} take-home ₪{s['est_takehome']:>6} | civic pool ₪{s['civic_pool']:>6}/mo")
